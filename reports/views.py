from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from core.models import Fundraiser, Transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io

@login_required
def export_transactions_excel(request, fundraiser_id=None):
    """
    Export transactions to Excel. If fundraiser_id is provided, export for that fundraiser.
    Otherwise, export all transactions.
    """
    if fundraiser_id:
        fundraiser = get_object_or_404(Fundraiser, id=fundraiser_id)
        transactions = Transaction.objects.filter(fundraiser=fundraiser)
        filename = f"PayDay_{fundraiser.slug}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    else:
        transactions = Transaction.objects.all()
        filename = f"PayDay_All_Transactions_{datetime.now().strftime('%Y%m%d')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Headers
    headers = ['ID', 'Date', 'Collecte', 'Donateur', 'Téléphone', 'Opérateur', 'Montant (FCFA)', 'Statut', 'Message']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for txn in transactions:
        ws.append([
            txn.id,
            txn.created_at.strftime("%d/%m/%Y %H:%M"),
            txn.fundraiser.title,
            txn.donor_name or "Anonyme",
            txn.donor_phone,
            txn.provider,
            txn.amount,
            txn.get_status_display(),
            txn.message or ""
        ])

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
